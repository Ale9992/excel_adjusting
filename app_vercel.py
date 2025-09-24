from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import pandas as pd
import tempfile
import os
import json
from typing import Optional
from solver_semplice import ExcelSolverSemplice as ExcelSolver

app = FastAPI(title="Excel Adjuster", description="Applicazione per correzione automatica di file Excel")

# Configurazione CORS per permettere richieste dal frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In produzione, specifica i domini esatti
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Excel Adjuster API", "status": "running"}

@app.post("/introspect")
async def introspect_excel(file: UploadFile = File(...)):
    """
    Analizza un file Excel e restituisce informazioni sui fogli e colonne
    """
    try:
        # Validazione del file
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File deve essere un Excel (.xlsx o .xls)")
        
        # Salva il file temporaneamente
        tmp_file_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                content = await file.read()
                tmp_file.write(content)
                tmp_file_path = tmp_file.name
            
            # Leggi il file Excel
            excel_file = pd.ExcelFile(tmp_file_path)
            sheets_info = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(tmp_file_path, sheet_name=sheet_name)
                    
                    # Identifica le colonne numeriche
                    numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
                    
                    # Prendi un campione dei dati
                    sample_data = df.head(5).to_dict('records') if len(df) > 0 else []
                    
                    sheets_info.append({
                        "name": sheet_name,
                        "rows": len(df),
                        "columns": list(df.columns),
                        "numeric_columns": numeric_columns,
                        "sample_data": sample_data
                    })
                except Exception as e:
                    sheets_info.append({
                        "name": sheet_name,
                        "error": str(e)
                    })
            
            return {
                "success": True,
                "sheets": sheets_info,
                "filename": file.filename
            }
            
        finally:
            # Pulisce il file temporaneo
            if tmp_file_path and os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nell'analisi del file: {str(e)}")

@app.post("/adjust")
async def adjust_excel(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    quantity_column: str = Form(...),
    price_column: str = Form(...),
    remaining_column: str = Form(...),
    target_total: float = Form(...),
    data_rows: int = Form(...)
):
    """
    Applica l'algoritmo di correzione al file Excel e restituisce il file modificato
    """
    try:
        # Validazione input
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File deve essere un Excel (.xlsx o .xls)")
        
        if target_total <= 0:
            raise HTTPException(status_code=400, detail="Il totale target deve essere positivo")
        
        # Salva il file temporaneamente
        tmp_file_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                content = await file.read()
                tmp_file.write(content)
                tmp_file_path = tmp_file.name
            
            # Crea il solver
            solver = ExcelSolver(
                file_path=tmp_file_path,
                sheet_name=sheet_name,
                quantity_column=quantity_column,
                price_column=price_column,
                remaining_column=remaining_column,
                target_total=target_total,
                data_rows=data_rows
            )
            
            # Esegui la correzione
            result = solver.adjust()
            
            if not result["success"]:
                raise HTTPException(status_code=500, detail=result.get("error", "Errore sconosciuto"))
            
            # Crea il file di output
            output_filename = f"adjusted_{file.filename}"
            output_path = os.path.join(tempfile.gettempdir(), output_filename)
            
            # Salva il file modificato preservando le formule originali
            from openpyxl import load_workbook
            
            # Carica il workbook originale
            wb = load_workbook(tmp_file_path)
            
            # Lavora sul foglio specificato
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Trova gli indici delle colonne
                quantity_col_idx = None
                price_col_idx = None
                remaining_col_idx = None
                
                # Cerca le colonne nell'header (prima riga)
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == solver.quantity_column:
                        quantity_col_idx = col_idx
                    elif cell.value == solver.price_column:
                        price_col_idx = col_idx
                    elif cell.value == solver.remaining_column:
                        remaining_col_idx = col_idx
                
                if quantity_col_idx and price_col_idx and remaining_col_idx:
                    # Aggiorna le colonne Quantità, Prezzo e Rimanenze
                    # Escludi la riga con la formula speciale se presente
                    max_data_row = ws.max_row
                    if hasattr(solver, 'special_formula_row') and solver.special_formula_row:
                        max_data_row = solver.special_formula_row - 1  # Escludi la riga con formula speciale
                        print(f"Esclusa riga {solver.special_formula_row} con formula speciale")
                    
                    # Aggiorna solo le righe che esistono nel DataFrame del solver
                    for i in range(len(solver.df)):
                        row_idx = i + 2  # Inizia dalla riga 2 (dopo l'header)
                        if row_idx <= max_data_row:
                            # Aggiorna quantità
                            ws.cell(row=row_idx, column=quantity_col_idx, value=solver.df.iloc[i][solver.quantity_column])
                            # Aggiorna prezzo
                            ws.cell(row=row_idx, column=price_col_idx, value=solver.df.iloc[i][solver.price_column])
                            # Aggiorna rimanenze (calcolate)
                            ws.cell(row=row_idx, column=remaining_col_idx, value=solver.df.iloc[i][solver.remaining_column])
                    
                    # Aggiorna la formula speciale se presente
                    if hasattr(solver, 'special_formula_row') and solver.special_formula_row:
                        special_row = solver.special_formula_row
                        # Aggiorna la formula speciale con il nuovo target
                        new_formula = f'=SUMIF(J2:J{max_data_row},">0") * ({int(target_total)} / SUMIF(J2:J{max_data_row},">0"))'
                        ws.cell(row=special_row, column=remaining_col_idx, value=new_formula)
                        print(f"Aggiornata formula speciale alla riga {special_row}")
            
            # Salva il workbook modificato
            wb.save(output_path)
            
            # Restituisci il file
            return FileResponse(
                output_path,
                filename=output_filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    "X-Original-Total": str(result["original_total"]),
                    "X-Final-Total": str(result["final_total"]),
                    "X-Target-Total": str(result["target_total"]),
                    "X-Precision": str(result.get("precision", 0))
                }
            )
            
        finally:
            # Pulisce i file temporanei
            if tmp_file_path and os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nella correzione del file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
