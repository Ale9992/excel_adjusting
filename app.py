from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import tempfile
import os
import json
from typing import Optional
from solver_semplice import ExcelSolverSemplice as ExcelSolver

def analyze_column_patterns(df, numeric_columns):
    """
    Analizza i pattern delle colonne numeriche per identificare automaticamente
    quantità, prezzi e rimanenze basandosi sui valori tipici
    """
    column_analysis = {}
    
    for col in numeric_columns:
        # Estrai i valori non nulli
        values = df[col].dropna()
        if len(values) == 0:
            continue
            
        # Calcola statistiche
        mean_val = values.mean()
        median_val = values.median()
        std_val = values.std()
        min_val = values.min()
        max_val = values.max()
        
        # Conta valori interi vs decimali
        integer_count = sum(1 for v in values if isinstance(v, (int, float)) and v == int(v))
        decimal_count = len(values) - integer_count
        
        # Analizza i pattern
        analysis = {
            'column': col,
            'mean': float(mean_val) if not pd.isna(mean_val) else 0,
            'median': float(median_val) if not pd.isna(median_val) else 0,
            'std': float(std_val) if not pd.isna(std_val) else 0,
            'min': float(min_val) if not pd.isna(min_val) else 0,
            'max': float(max_val) if not pd.isna(max_val) else 0,
            'integer_ratio': integer_count / len(values) if len(values) > 0 else 0,
            'decimal_ratio': decimal_count / len(values) if len(values) > 0 else 0,
            'likely_type': 'unknown'
        }
        
        # Euristiche per identificare il tipo di colonna
        
        # QUANTITÀ: valori interi, range tipico 1-1000, bassa varianza
        if (analysis['integer_ratio'] > 0.8 and 
            analysis['mean'] > 0 and analysis['mean'] < 1000 and
            analysis['std'] < analysis['mean'] * 0.5):
            analysis['likely_type'] = 'quantity'
            analysis['confidence'] = 0.8
        
        # PREZZI: valori decimali, range tipico 0.01-1000, alta varianza
        elif (analysis['decimal_ratio'] > 0.4 and 
              analysis['mean'] > 0.01 and analysis['mean'] < 10000 and
              analysis['std'] > analysis['mean'] * 0.2):
            analysis['likely_type'] = 'price'
            analysis['confidence'] = 0.7
        
        # RIMANENZE/TOTALI: valori decimali, range ampio, alta varianza
        elif (analysis['mean'] > 10 and 
              analysis['std'] > analysis['mean'] * 0.4 and
              analysis['max'] > analysis['mean'] * 2):
            analysis['likely_type'] = 'remaining'
            analysis['confidence'] = 0.6
        
        # VALORI PICCOLI: potrebbero essere quantità
        elif (analysis['mean'] < 100 and analysis['integer_ratio'] > 0.6):
            analysis['likely_type'] = 'quantity'
            analysis['confidence'] = 0.5
        
        # VALORI GRANDI: potrebbero essere totali
        elif (analysis['mean'] > 100):
            analysis['likely_type'] = 'remaining'
            analysis['confidence'] = 0.5
        
        column_analysis[col] = analysis
    
    return column_analysis

app = FastAPI(title="Excel Adjuster", description="Applicazione per correzione automatica di file Excel")

# Configurazione CORS per permettere richieste dal frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve file statici (CSS, JS, immagini)
app.mount("/assets", StaticFiles(directory="assets"), name="assets")

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve la pagina principale dell'applicazione"""
    try:
        with open("index.html", "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read(), status_code=200)
    except FileNotFoundError:
        return HTMLResponse(content="<h1>File index.html non trovato</h1>", status_code=404)

@app.get("/app.js")
async def serve_js():
    """Serve il file JavaScript dell'applicazione"""
    try:
        with open("app.js", "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read(), media_type="application/javascript")
    except FileNotFoundError:
        return HTMLResponse(content="// File app.js non trovato", status_code=404)

@app.get("/api/status")
async def api_status():
    """Endpoint per verificare lo stato dell'API"""
    return {
        "message": "Excel Adjuster API - Backend attivo",
        "status": "running",
        "version": "1.0.0",
        "environment": "production" if os.getenv("RENDER") else "development"
    }

@app.get("/app.js")
async def serve_js():
    """Serve il file JavaScript"""
    try:
        with open("app.js", "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read(), status_code=200, media_type="application/javascript")
    except FileNotFoundError:
        return HTMLResponse(content="// File app.js non trovato", status_code=404)

@app.post("/introspect")
async def introspect_excel(file: UploadFile = File(...)):
    """
    Analizza un file Excel e restituisce informazioni sui fogli e colonne disponibili
    """
    try:
        # Verifica che sia un file Excel
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Il file deve essere un Excel (.xlsx o .xls)")
        
        # Salva temporaneamente il file con l'estensione corretta
        file_extension = '.xlsx' if file.filename.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        try:
            # Legge il file Excel
            excel_file = pd.ExcelFile(tmp_file_path)
            
            # Estrae informazioni sui fogli
            sheets_info = {}
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(tmp_file_path, sheet_name=sheet_name)
                
                # Filtra solo le colonne numeriche e pulisce i dati
                numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
                
                # Pulisce i dati per evitare valori inf/nan
                clean_df = df.copy()
                for col in numeric_columns:
                    # Sostituisce inf e nan con 0
                    clean_df[col] = clean_df[col].replace([float('inf'), float('-inf')], 0)
                    clean_df[col] = clean_df[col].fillna(0)
                
                # Analizza i pattern delle colonne per identificazione automatica
                column_analysis = analyze_column_patterns(clean_df, numeric_columns)
                
                # Prepara i dati di esempio in modo sicuro
                sample_data = []
                if len(clean_df) > 0:
                    sample_df = clean_df.head(3)
                    for _, row in sample_df.iterrows():
                        sample_row = {}
                        for col in numeric_columns:
                            value = row[col]
                            # Converte in float sicuro per JSON
                            if pd.isna(value) or value == float('inf') or value == float('-inf'):
                                sample_row[col] = 0.0
                            else:
                                sample_row[col] = float(value)
                        sample_data.append(sample_row)
                
                # Identifica automaticamente le colonne più probabili
                suggested_columns = {
                    "quantity": None,
                    "price": None,
                    "remaining": None
                }
                
                # Trova le colonne con la confidenza più alta per ogni tipo
                for col, analysis in column_analysis.items():
                    col_type = analysis.get('likely_type', 'unknown')
                    confidence = analysis.get('confidence', 0)
                    
                    if col_type in suggested_columns:
                        current_confidence = 0
                        if suggested_columns[col_type]:
                            current_confidence = column_analysis[suggested_columns[col_type]].get('confidence', 0)
                        
                        if confidence > current_confidence:
                            suggested_columns[col_type] = col
                
                sheets_info[sheet_name] = {
                    "columns": numeric_columns,
                    "row_count": len(df),
                    "sample_data": sample_data,
                    "column_analysis": column_analysis,
                    "suggested_columns": suggested_columns
                }
            
            return {
                "success": True,
                "sheets": sheets_info,
                "filename": file.filename
            }
            
        finally:
            # Pulisce il file temporaneo
            os.unlink(tmp_file_path)
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nell'analisi del file: {str(e)}")

@app.post("/adjust")
async def adjust_excel(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    quantity_column: str = Form(...),
    price_column: str = Form(...),
    remaining_column: str = Form(...),
    target_total: float = Form(...),
    data_rows: int = Form(...)
):
    """
    Applica l'algoritmo di correzione al file Excel e restituisce il file modificato
    """
    try:
        # Validazione input
        if target_total <= 0:
            raise HTTPException(status_code=400, detail="Il totale target deve essere maggiore di 0")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Il file deve essere un Excel (.xlsx o .xls)")
        
        # Salva temporaneamente il file con l'estensione corretta
        file_extension = '.xlsx' if file.filename.endswith('.xlsx') else '.xls'
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        try:
            print(f"Creazione solver con parametri:")
            print(f"  file_path: {tmp_file_path}")
            print(f"  sheet_name: {sheet_name}")
            print(f"  quantity_column: {quantity_column}")
            print(f"  price_column: {price_column}")
            print(f"  remaining_column: {remaining_column}")
            print(f"  target_total: {target_total}")
            print(f"  data_rows: {data_rows}")
            
            # Inizializza il solver con la nuova logica intelligente
            solver = ExcelSolver(
                file_path=tmp_file_path,
                sheet_name=sheet_name,
                quantity_column=quantity_column,
                price_column=price_column,
                remaining_column=remaining_column,
                target_total=target_total,
                data_rows=data_rows
            )
            
            print("Solver creato con successo")
            
            # Esegue la correzione
            result = solver.adjust()
            
            if not result["success"]:
                raise HTTPException(status_code=400, detail=result["error"])
            
            # Crea il file di output
            # Se il file originale era .xls, salva come .xlsx (conversione automatica)
            if file_extension == '.xls':
                output_filename = f"adjusted_{file.filename.replace('.xls', '.xlsx')}"
            else:
                output_filename = f"adjusted_{file.filename}"
            output_path = os.path.join(tempfile.gettempdir(), output_filename)
            
            # Salva il file modificato preservando le formule originali
            from openpyxl import load_workbook
            import pandas as pd
            
            # Se il file è .xls, convertilo in .xlsx per l'elaborazione
            if file_extension == '.xls':
                # Leggi con pandas (supporta .xls) e salva come .xlsx
                xlsx_path = tmp_file_path.replace('.xls', '.xlsx')
                df_dict = pd.read_excel(tmp_file_path, sheet_name=None, engine='xlrd')
                
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    for sheet_name_df, df in df_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name_df, index=False)
                
                # Usa il file convertito per l'elaborazione
                wb = load_workbook(xlsx_path)
            else:
                # File .xlsx, carica direttamente
                wb = load_workbook(tmp_file_path)
            
            # Lavora sul foglio specificato
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Trova gli indici delle colonne
                quantity_col_idx = None
                price_col_idx = None
                remaining_col_idx = None
                
                # Cerca le colonne nell'header (prima riga)
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == solver.quantity_column:
                        quantity_col_idx = col_idx
                    elif cell.value == solver.price_column:
                        price_col_idx = col_idx
                    elif cell.value == solver.remaining_column:
                        remaining_col_idx = col_idx
                
                if quantity_col_idx and price_col_idx and remaining_col_idx:
                    # Aggiorna le colonne Quantità, Prezzo e Rimanenze
                    # Escludi la riga con la formula speciale se presente
                    max_data_row = ws.max_row
                    if hasattr(solver, 'special_formula_row') and solver.special_formula_row:
                        max_data_row = solver.special_formula_row - 1  # Escludi la riga con formula speciale
                        print(f"Esclusa riga {solver.special_formula_row} con formula speciale")
                    
                    # Aggiorna solo le righe che esistono nel DataFrame del solver
                    for i in range(len(solver.df)):
                        row_idx = i + 2  # Inizia dalla riga 2 (dopo l'header)
                        if row_idx <= max_data_row:
                            # Aggiorna quantità
                            ws.cell(row=row_idx, column=quantity_col_idx, value=solver.df.iloc[i][solver.quantity_column])
                            # Aggiorna prezzo
                            ws.cell(row=row_idx, column=price_col_idx, value=solver.df.iloc[i][solver.price_column])
                            # Aggiorna rimanenze (calcolate)
                            ws.cell(row=row_idx, column=remaining_col_idx, value=solver.df.iloc[i][solver.remaining_column])
                    
                    # Aggiorna la formula speciale se presente
                    if hasattr(solver, 'special_formula_row') and solver.special_formula_row:
                        special_row = solver.special_formula_row
                        # Aggiorna la formula speciale con il nuovo target
                        new_formula = f'=SUMIF(J2:J{max_data_row},">0") * ({int(target_total)} / SUMIF(J2:J{max_data_row},">0"))'
                        ws.cell(row=special_row, column=10, value=new_formula)
                        print(f"Aggiornata formula speciale alla riga {special_row}: {new_formula}")
            
            # Forza il ricalcolo delle formule
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True
            
            # Forza il ricalcolo di tutte le formule nel foglio
            for sheet in wb.worksheets:
                sheet.calculate_dimension()
            
            # Salva il workbook modificato
            wb.save(output_path)
            
            # Salva le statistiche in un file temporaneo per il frontend
            stats_filename = f"stats_{file.filename}.json"
            stats_path = os.path.join(tempfile.gettempdir(), stats_filename)
            
            # Converte i tipi NumPy in tipi Python nativi per la serializzazione JSON
            serializable_result = {
                "success": result["success"],
                "message": result["message"],
                "original_total": float(result["original_total"]),
                "final_total": float(result["final_total"]),
                "target_total": float(result["target_total"])
            }
            
            with open(stats_path, 'w') as f:
                import json
                json.dump(serializable_result, f)
            
            # Aggiunge le statistiche agli header della risposta
            headers = {
                'X-Original-Total': str(result.get('original_total', 0)),
                'X-Target-Total': str(result.get('target_total', 0)),
                'X-Final-Total': str(result.get('final_total', 0)),
                'X-Difference': str(result.get('difference', 0)),
                'X-Rows-Processed': str(result.get('rows_processed', 0))
            }
            
            return FileResponse(
                path=output_path,
                filename=output_filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers=headers
            )
            
        finally:
            # Pulisce il file temporaneo di input
            os.unlink(tmp_file_path)
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore nella correzione del file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
